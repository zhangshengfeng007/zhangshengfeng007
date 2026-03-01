#!/usr/bin/env python3
"""Offline grayscale statistics tool for image files.

Supports both CLI and GUI mode.
"""

import argparse
import json
import math
import os
import sys
from pathlib import Path
from typing import Dict, List, Sequence, Set


def _hide_console_window() -> None:
    if not sys.platform.startswith("win"):
        return
    if os.getenv("IMAGE_GRAY_CHECK_SHOW_CONSOLE"):
        return
    try:
        import ctypes

        kernel32 = ctypes.windll.kernel32
        user32 = ctypes.windll.user32
        hwnd = kernel32.GetConsoleWindow()
        if hwnd:
            user32.ShowWindow(hwnd, 0)
            kernel32.FreeConsole()
    except Exception:
        pass


def _import_pil_image():
    try:
        from PIL import Image  # type: ignore

        return Image
    except Exception:
        repo_root = Path(__file__).resolve().parents[1]
        local_pil_path = repo_root / "build" / "scripts"
        if local_pil_path.is_dir():
            sys.path.insert(0, str(local_pil_path))
            from PIL import Image  # type: ignore

            return Image
        raise


Image = _import_pil_image()


DEFAULT_EXTENSIONS = {
    ".png",
    ".jpg",
    ".jpeg",
    ".bmp",
    ".gif",
    ".webp",
    ".tif",
    ".tiff",
}


def _get_excel_output_path() -> Path:
    if getattr(sys, "frozen", False):
        base_dir = Path(sys.executable).resolve().parent
    else:
        base_dir = Path(__file__).resolve().parent
    return base_dir / "image_gray_check_results.xlsx"


def _export_results_to_excel(
    stats_list: Sequence[Dict[str, object]],
    output_path: Path,
    timestamp: str,
) -> str:
    if not stats_list:
        return "无可导出的结果。"
    try:
        import openpyxl
    except Exception as exc:
        return f"Excel 导出失败: {exc}"

    headers = [
        "Name",
        "TestResult",
        "Timestamp",
        "Path",
        "Width",
        "Height",
        "GrayMean",
        "GrayStddev",
        "GrayMin",
        "GrayMax",
        "DarkRatioLt64",
        "DarkRatioLt128",
        "RulePassPixels",
        "RuleFailPixels",
    ]

    existing_order: List[str] = []
    existing_records: Dict[str, Dict[str, object]] = {}

    if output_path.exists():
        wb = openpyxl.load_workbook(output_path)
        if "GrayCheck" in wb.sheetnames:
            ws_old = wb["GrayCheck"]
        else:
            ws_old = wb.active

        header_to_col: Dict[str, int] = {}
        for col in range(1, ws_old.max_column + 1):
            value = ws_old.cell(row=1, column=col).value
            if isinstance(value, str) and value:
                header_to_col[value] = col

        name_col = header_to_col.get("Name", 1)
        for row in range(2, ws_old.max_row + 1):
            raw_name = ws_old.cell(row=row, column=name_col).value
            if not raw_name:
                continue
            name = str(raw_name)
            if name not in existing_records:
                existing_order.append(name)
            record: Dict[str, object] = existing_records.get(name, {})
            for header in headers:
                col = header_to_col.get(header)
                if col is not None:
                    record[header] = ws_old.cell(row=row, column=col).value
            existing_records[name] = record

        wb.remove(ws_old)
        ws = wb.create_sheet("GrayCheck", 0)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GrayCheck"

    added = 0
    updated = 0

    for stats in stats_list:
        name = Path(str(stats["path"])).name
        if name not in existing_records:
            existing_order.append(name)
            added += 1
        else:
            updated += 1

        record = existing_records.get(name, {})
        record.update(
            {
                "Name": name,
                "TestResult": stats["test_result"],
                "Timestamp": timestamp,
                "Path": str(stats["path"]),
                "Width": stats["width"],
                "Height": stats["height"],
                "GrayMean": stats["gray_mean"],
                "GrayStddev": stats["gray_stddev"],
                "GrayMin": stats["gray_min"],
                "GrayMax": stats["gray_max"],
                "DarkRatioLt64": stats["dark_ratio_lt_64_percent"],
                "DarkRatioLt128": stats["dark_ratio_lt_128_percent"],
                "RulePassPixels": stats["rule_pass_pixels"],
                "RuleFailPixels": stats["rule_fail_pixels"],
            }
        )
        existing_records[name] = record

    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)

    try:
        from openpyxl.styles import PatternFill
    except Exception:
        PatternFill = None  # type: ignore

    pass_fill = None
    fail_fill = None
    if PatternFill is not None:
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    result_col_index = headers.index("TestResult") + 1

    row = 2
    for name in existing_order:
        record = existing_records.get(name, {"Name": name})
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=record.get(header))
            if col == result_col_index and pass_fill is not None and fail_fill is not None:
                result_value = str(record.get("TestResult", "")).upper()
                if result_value == "PASS":
                    cell.fill = pass_fill
                elif result_value == "FAIL":
                    cell.fill = fail_fill
        row += 1

    try:
        wb.save(output_path)
    except Exception as exc:
        return f"Excel 保存失败: {exc}"

    return (
        f"Excel 已更新: 新增 {added} 条, 覆盖 {updated} 条。"
        f" 文件: {output_path}"
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Check grayscale values of image files. "
            "For RGB pixels, gray value uses max(R,G,B)."
        )
    )
    parser.add_argument(
        "inputs",
        nargs="*",
        help="Image file(s) or directory path(s). If empty, GUI mode starts.",
    )
    parser.add_argument(
        "-r",
        "--recursive",
        action="store_true",
        help="Recursively scan directories",
    )
    parser.add_argument(
        "--ignore-transparent",
        action="store_true",
        help="Ignore pixels whose alpha is 0",
    )
    parser.add_argument(
        "--hist",
        action="store_true",
        help="Print histogram data",
    )
    parser.add_argument(
        "--hist-all",
        action="store_true",
        help="Print all 0-255 histogram bins (works with --hist)",
    )
    parser.add_argument(
        "--json",
        action="store_true",
        help="Print machine-readable JSON output",
    )
    parser.add_argument(
        "--gui",
        action="store_true",
        help="Force GUI mode",
    )
    return parser.parse_args()


def collect_images(inputs: Sequence[str], recursive: bool) -> List[Path]:
    files: List[Path] = []
    seen = set()

    for raw in inputs:
        path = Path(raw)
        if path.is_file():
            resolved = path.resolve()
            if resolved not in seen:
                files.append(resolved)
                seen.add(resolved)
            continue

        if path.is_dir():
            iterator = path.rglob("*") if recursive else path.glob("*")
            for child in iterator:
                if not child.is_file():
                    continue
                if child.suffix.lower() not in DEFAULT_EXTENSIONS:
                    continue
                resolved = child.resolve()
                if resolved not in seen:
                    files.append(resolved)
                    seen.add(resolved)
            continue

        raise FileNotFoundError(f"Input not found: {raw}")

    files.sort()
    return files


def quantile_from_histogram(hist: Sequence[int], total: int, q: float) -> int:
    if total <= 0:
        return 0
    target = int(math.floor((total - 1) * q))
    cumulative = 0
    for value, count in enumerate(hist):
        cumulative += count
        if cumulative > target:
            return value
    return 255


def analyze_image(path: Path, ignore_transparent: bool) -> Dict[str, object]:
    with Image.open(path) as img:
        width, height = img.size
        mode = img.mode
        total_pixels = width * height
        rule_pass = 0
        rule_fail = 0

        if mode in {"1", "L"}:
            data = img.convert("L").getdata()
            hist = [0] * 256
            valid = 0
            gray_sum = 0
            gray_sq_sum = 0
            min_gray = 255
            max_gray = 0

            for gray in data:
                gray = int(gray)
                if gray == 0:
                    # Pure black means "no data", skip it.
                    continue
                hist[gray] += 1
                valid += 1
                if gray > 153:
                    rule_pass += 1
                else:
                    rule_fail += 1
                gray_sum += gray
                gray_sq_sum += gray * gray
                min_gray = min(min_gray, gray)
                max_gray = max(max_gray, gray)
        else:
            rgba = img.convert("RGBA")
            hist = [0] * 256
            valid = 0
            gray_sum = 0
            gray_sq_sum = 0
            min_gray = 255
            max_gray = 0

            for r, g, b, a in rgba.getdata():
                if r == 0 and g == 0 and b == 0:
                    # Pure black means "no data", skip it.
                    continue
                if ignore_transparent and a == 0:
                    continue
                # gray<T is equivalent to R<T and G<T and B<T.
                gray = max(r, g, b)
                hist[gray] += 1
                valid += 1
                if r == g == b:
                    pixel_pass = r > 153
                else:
                    pixel_pass = gray >= 128
                if pixel_pass:
                    rule_pass += 1
                else:
                    rule_fail += 1
                gray_sum += gray
                gray_sq_sum += gray * gray
                min_gray = min(min_gray, gray)
                max_gray = max(max_gray, gray)

        if valid == 0:
            raise ValueError("No valid pixels after filtering")

        mean = gray_sum / valid
        variance = gray_sq_sum / valid - mean * mean
        stddev = math.sqrt(max(0.0, variance))
        dark_lt_64 = sum(hist[:64]) * 100.0 / valid
        dark_lt_128 = sum(hist[:128]) * 100.0 / valid
        dark_le_64 = sum(hist[:65]) * 100.0 / valid
        dark_le_128 = sum(hist[:129]) * 100.0 / valid
        used_ratio = valid * 100.0 / total_pixels if total_pixels else 0.0

        return {
            "path": str(path),
            "width": width,
            "height": height,
            "mode": mode,
            "total_pixels": total_pixels,
            "valid_pixels": valid,
            "used_ratio_percent": round(used_ratio, 4),
            "gray_mean": round(mean, 4),
            "gray_stddev": round(stddev, 4),
            "gray_min": min_gray,
            "gray_max": max_gray,
            "gray_p10": quantile_from_histogram(hist, valid, 0.10),
            "gray_median": quantile_from_histogram(hist, valid, 0.50),
            "gray_p90": quantile_from_histogram(hist, valid, 0.90),
            "dark_ratio_lt_64_percent": round(dark_lt_64, 4),
            "dark_ratio_lt_128_percent": round(dark_lt_128, 4),
            "dark_ratio_le_64_percent": round(dark_le_64, 4),
            "dark_ratio_le_128_percent": round(dark_le_128, 4),
            "rule_pass_pixels": rule_pass,
            "rule_fail_pixels": rule_fail,
            "test_result": "PASS" if rule_fail == 0 else "FAIL",
            "histogram": hist,
        }


def format_text(stats: Dict[str, object], show_hist: bool, hist_all: bool) -> str:
    lines = [
        f"Image: {stats['path']}",
        f"Size: {stats['width']}x{stats['height']}  mode={stats['mode']}",
        (
            "Pixels used: "
            f"{stats['valid_pixels']}/{stats['total_pixels']} "
            f"({stats['used_ratio_percent']:.2f}%)"
        ),
        f"Gray mean/stddev: {stats['gray_mean']:.4f} / {stats['gray_stddev']:.4f}",
        f"Gray min/max: {stats['gray_min']} / {stats['gray_max']}",
        (
            "Gray p10/median/p90: "
            f"{stats['gray_p10']} / {stats['gray_median']} / {stats['gray_p90']}"
        ),
        (
            "Dark ratio <64 / <128: "
            f"{stats['dark_ratio_lt_64_percent']:.2f}% / "
            f"{stats['dark_ratio_lt_128_percent']:.2f}%"
        ),
        f"Test result: {stats['test_result']}",
        (
            "Rule pass/fail pixels: "
            f"{stats['rule_pass_pixels']} / {stats['rule_fail_pixels']}"
        ),
        "Rule: gray=max(R,G,B), pure black pixels are ignored.",
    ]

    if show_hist:
        lines.append("Histogram (gray_value:count ratio%):")
        hist = stats["histogram"]
        valid = int(stats["valid_pixels"])
        for value, count in enumerate(hist):
            if not hist_all and count == 0:
                continue
            ratio = (count * 100.0 / valid) if valid else 0.0
            lines.append(f"  {value:3d}:{count:8d} {ratio:8.4f}%")

    return "\n".join(lines)


def format_text_cn(stats: Dict[str, object], show_hist: bool, hist_all: bool) -> str:
    display_result = "pass" if stats["test_result"] == "PASS" else "fail"
    lines = [
        f"图片: {stats['path']}",
        f"尺寸: {stats['width']}x{stats['height']}  模式={stats['mode']}",
        (
            "参与统计像素: "
            f"{stats['valid_pixels']}/{stats['total_pixels']} "
            f"({stats['used_ratio_percent']:.2f}%)"
        ),
        f"灰阶均值/标准差: {stats['gray_mean']:.4f} / {stats['gray_stddev']:.4f}",
        f"灰阶最小/最大: {stats['gray_min']} / {stats['gray_max']}",
        (
            "灰阶 p10/中位/p90: "
            f"{stats['gray_p10']} / {stats['gray_median']} / {stats['gray_p90']}"
        ),
        (
            "暗像素占比 <64 / <128: "
            f"{stats['dark_ratio_lt_64_percent']:.2f}% / "
            f"{stats['dark_ratio_lt_128_percent']:.2f}%"
        ),
        f"测试结果: {display_result}",
        (
            "规则通过/失败像素: "
            f"{stats['rule_pass_pixels']} / {stats['rule_fail_pixels']}"
        ),
        "说明: gray=max(R,G,B)，且已自动忽略纯黑像素 (R=G=B=0)",
        "规则: 非灰像素需至少一个通道>=128；灰色像素需 R=G=B>153",
    ]

    if show_hist:
        lines.append("直方图 (灰阶值:数量 占比):")
        hist = stats["histogram"]
        valid = int(stats["valid_pixels"])
        for value, count in enumerate(hist):
            if not hist_all and count == 0:
                continue
            ratio = (count * 100.0 / valid) if valid else 0.0
            lines.append(f"  {value:3d}:{count:8d} {ratio:8.4f}%")

    return "\n".join(lines)


def run_gui(
    initial_inputs: Sequence[str],
    recursive: bool,
    ignore_transparent: bool,
    show_hist: bool,
    hist_all: bool,
) -> int:
    try:
        import tkinter as tk
        from datetime import datetime
        from tkinter import filedialog, messagebox, ttk
        from tkinter.scrolledtext import ScrolledText
    except Exception as exc:
        print(f"GUI不可用: {exc}", file=sys.stderr)
        return 2

    image_tk_module = None
    try:
        from PIL import ImageTk  # type: ignore

        image_tk_module = ImageTk
    except Exception:
        image_tk_module = None

    class GrayCheckApp:
        def __init__(self, root: "tk.Tk") -> None:
            self.root = root
            self.selected_paths: List[str] = []
            self.selected_set: Set[str] = set()
            self.analysis_cache: Dict[str, Dict[str, object]] = {}

            self.ignore_transparent_var = tk.BooleanVar(value=ignore_transparent)
            self.show_hist_var = tk.BooleanVar(value=show_hist)
            self.hist_all_var = tk.BooleanVar(value=hist_all)
            self.status_var = tk.StringVar(
                value="就绪：请先选择图片，然后点击“开始分析”。"
            )
            self.preview_meta_var = tk.StringVar(value="请选择左侧图片查看缩略图。")
            self.result_var = tk.StringVar(value="未分析")
            self.result_detail_var = tk.StringVar(value="等待分析")
            self.preview_photo = None
            self.image_tk_module = image_tk_module

            self.colors = {
                "bg": "#e8eef5",
                "card": "#ffffff",
                "header": "#0b5ed7",
                "header_subtitle": "#dce8ff",
                "text": "#1f2a37",
                "muted": "#66788a",
                "accent": "#0f7bff",
                "accent_hover": "#0565d8",
                "success": "#2f9e44",
                "warning": "#f08c00",
                "danger": "#d9480f",
                "output_bg": "#0f172a",
                "output_fg": "#d8e4ff",
                "result_bg": "#f1f5f9",
                "result_border": "#d0d8e5",
            }

            self._configure_style()
            self._build_layout()
            self._set_result_status("未分析", "等待分析")
            self._load_initial_inputs(initial_inputs, recursive)

        def _configure_style(self) -> None:
            self.root.title("图片灰阶检查工具")
            self.root.geometry("1120x760")
            self.root.minsize(960, 640)
            self.root.configure(bg=self.colors["bg"])

            style = ttk.Style(self.root)
            try:
                style.theme_use("clam")
            except tk.TclError:
                pass

            style.configure("Main.TFrame", background=self.colors["bg"])
            style.configure("Card.TFrame", background=self.colors["card"])
            style.configure("Header.TFrame", background=self.colors["header"])

            style.configure(
                "HeaderTitle.TLabel",
                background=self.colors["header"],
                foreground="#ffffff",
                font=("Segoe UI", 16, "bold"),
            )
            style.configure(
                "HeaderSubtitle.TLabel",
                background=self.colors["header"],
                foreground=self.colors["header_subtitle"],
                font=("Segoe UI", 10),
            )
            style.configure(
                "CardTitle.TLabel",
                background=self.colors["card"],
                foreground=self.colors["text"],
                font=("Segoe UI", 11, "bold"),
            )
            style.configure(
                "Hint.TLabel",
                background=self.colors["card"],
                foreground=self.colors["muted"],
                font=("Segoe UI", 9),
            )
            style.configure(
                "Status.TLabel",
                background=self.colors["bg"],
                foreground=self.colors["text"],
                font=("Segoe UI", 9, "bold"),
            )

            style.configure(
                "Accent.TButton",
                padding=(14, 8),
                background=self.colors["accent"],
                foreground="#ffffff",
                font=("Segoe UI", 9, "bold"),
                borderwidth=0,
                relief="flat",
            )
            style.map(
                "Accent.TButton",
                background=[("active", self.colors["accent_hover"])],
                foreground=[("disabled", "#c7d2e8")],
            )

            style.configure(
                "Subtle.TButton",
                padding=(12, 8),
                background="#f1f4f9",
                foreground=self.colors["text"],
                font=("Segoe UI", 9),
                borderwidth=1,
                relief="flat",
            )
            style.map(
                "Subtle.TButton",
                background=[("active", "#e5ebf4")],
                foreground=[("disabled", "#9aa7b6")],
            )

            style.configure(
                "Option.TCheckbutton",
                background=self.colors["card"],
                foreground=self.colors["text"],
                font=("Segoe UI", 9),
            )

        def _build_layout(self) -> None:
            self.root.columnconfigure(0, weight=1)
            self.root.rowconfigure(1, weight=1)

            header = ttk.Frame(self.root, style="Header.TFrame", padding=(18, 14))
            header.grid(row=0, column=0, sticky="ew")
            header.columnconfigure(0, weight=1)

            ttk.Label(
                header,
                text="图片灰阶检查",
                style="HeaderTitle.TLabel",
            ).grid(row=0, column=0, sticky="w")
            ttk.Label(
                header,
                text=(
                    "支持多次选择图片、重复分析，结果会保留在窗口中便于对比。"
                ),
                style="HeaderSubtitle.TLabel",
            ).grid(row=1, column=0, sticky="w", pady=(2, 0))

            body = ttk.Frame(self.root, style="Main.TFrame", padding=(14, 12))
            body.grid(row=1, column=0, sticky="nsew")
            body.columnconfigure(0, weight=3)
            body.columnconfigure(1, weight=5)
            body.rowconfigure(1, weight=1)

            control_card = ttk.Frame(body, style="Card.TFrame", padding=12)
            control_card.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 10))
            control_card.columnconfigure(6, weight=1)

            ttk.Button(
                control_card,
                text="选择图片",
                style="Accent.TButton",
                command=self.select_images,
            ).grid(row=0, column=0, padx=(0, 8), pady=2)

            ttk.Button(
                control_card,
                text="开始分析",
                style="Accent.TButton",
                command=self.analyze_selected,
            ).grid(row=0, column=1, padx=(0, 8), pady=2)

            ttk.Button(
                control_card,
                text="移除选中",
                style="Subtle.TButton",
                command=self.remove_selected_items,
            ).grid(row=0, column=2, padx=(0, 8), pady=2)

            ttk.Button(
                control_card,
                text="清空列表",
                style="Subtle.TButton",
                command=self.clear_files,
            ).grid(row=0, column=3, padx=(0, 8), pady=2)

            ttk.Button(
                control_card,
                text="清空结果",
                style="Subtle.TButton",
                command=self.clear_output,
            ).grid(row=0, column=4, padx=(0, 8), pady=2)

            ttk.Checkbutton(
                control_card,
                text="忽略全透明像素",
                variable=self.ignore_transparent_var,
                style="Option.TCheckbutton",
            ).grid(row=0, column=5, padx=(8, 10), pady=2, sticky="w")

            hist_frame = ttk.Frame(control_card, style="Card.TFrame")
            hist_frame.grid(row=0, column=6, sticky="e")
            ttk.Checkbutton(
                hist_frame,
                text="显示直方图",
                variable=self.show_hist_var,
                style="Option.TCheckbutton",
                command=self._sync_hist_all_state,
            ).grid(row=0, column=0, sticky="e")

            self.hist_all_check = ttk.Checkbutton(
                hist_frame,
                text="显示全部灰阶桶",
                variable=self.hist_all_var,
                style="Option.TCheckbutton",
            )
            self.hist_all_check.grid(row=0, column=1, padx=(10, 0), sticky="e")
            self._sync_hist_all_state()

            files_card = ttk.Frame(body, style="Card.TFrame", padding=10)
            files_card.grid(row=1, column=0, sticky="nsew", padx=(0, 8))
            files_card.columnconfigure(0, weight=1)
            files_card.rowconfigure(2, weight=3)
            files_card.rowconfigure(4, weight=2)

            ttk.Label(
                files_card, text="已选图片文件", style="CardTitle.TLabel"
            ).grid(row=0, column=0, sticky="w")
            ttk.Label(
                files_card,
                text="提示：可重复点击“选择图片”追加文件，重复文件会自动跳过。",
                style="Hint.TLabel",
            ).grid(row=1, column=0, sticky="w", pady=(0, 6))

            list_frame = tk.Frame(files_card, bg=self.colors["card"])
            list_frame.grid(row=2, column=0, sticky="nsew")
            list_frame.columnconfigure(0, weight=1)
            list_frame.rowconfigure(0, weight=1)

            self.file_list = tk.Listbox(
                list_frame,
                selectmode="extended",
                activestyle="none",
                bg="#f6f9ff",
                fg=self.colors["text"],
                selectbackground="#cce0ff",
                selectforeground=self.colors["text"],
                highlightthickness=1,
                highlightbackground="#d0d8e5",
                relief="flat",
                font=("Consolas", 9),
            )
            self.file_list.grid(row=0, column=0, sticky="nsew")
            self.file_list.bind("<<ListboxSelect>>", self._on_file_selection)

            file_scroll = tk.Scrollbar(list_frame, command=self.file_list.yview)
            file_scroll.grid(row=0, column=1, sticky="ns")
            self.file_list.configure(yscrollcommand=file_scroll.set)

            ttk.Label(files_card, text="图片预览", style="CardTitle.TLabel").grid(
                row=3, column=0, sticky="w", pady=(8, 4)
            )

            preview_frame = tk.Frame(
                files_card,
                bg="#f6f9ff",
                highlightthickness=1,
                highlightbackground="#d0d8e5",
            )
            preview_frame.grid(row=4, column=0, sticky="nsew")
            preview_frame.columnconfigure(0, weight=1)
            preview_frame.rowconfigure(0, weight=1)

            self.preview_label = tk.Label(
                preview_frame,
                bg="#f6f9ff",
                fg=self.colors["muted"],
                text="暂无预览",
                font=("Segoe UI", 10),
            )
            self.preview_label.grid(row=0, column=0, sticky="nsew")

            preview_meta = tk.Label(
                preview_frame,
                bg="#f6f9ff",
                fg=self.colors["muted"],
                textvariable=self.preview_meta_var,
                anchor="w",
                justify="left",
                padx=8,
                pady=6,
                font=("Segoe UI", 9),
            )
            preview_meta.grid(row=1, column=0, sticky="ew")

            output_card = ttk.Frame(body, style="Card.TFrame", padding=10)
            output_card.grid(row=1, column=1, sticky="nsew")
            output_card.columnconfigure(0, weight=1)
            output_card.rowconfigure(2, weight=1)

            ttk.Label(output_card, text="分析结果", style="CardTitle.TLabel").grid(
                row=0, column=0, sticky="w", pady=(0, 6)
            )

            result_box = tk.Frame(
                output_card,
                bg=self.colors["result_bg"],
                highlightthickness=1,
                highlightbackground=self.colors["result_border"],
                width=220,
                height=80,
            )
            result_box.grid(row=1, column=0, sticky="w", pady=(0, 10))
            result_box.grid_propagate(False)

            self.result_box = result_box
            self.result_title = tk.Label(
                result_box,
                text="测试结果",
                bg=self.colors["result_bg"],
                fg=self.colors["muted"],
                font=("Segoe UI", 9, "bold"),
                anchor="w",
            )
            self.result_title.pack(anchor="w", padx=10, pady=(8, 0))
            self.result_value = tk.Label(
                result_box,
                textvariable=self.result_var,
                bg=self.colors["result_bg"],
                fg=self.colors["text"],
                font=("Segoe UI", 16, "bold"),
                anchor="w",
            )
            self.result_value.pack(anchor="w", padx=10, pady=(2, 0))
            self.result_detail = tk.Label(
                result_box,
                textvariable=self.result_detail_var,
                bg=self.colors["result_bg"],
                fg=self.colors["muted"],
                font=("Segoe UI", 9),
                anchor="w",
            )
            self.result_detail.pack(anchor="w", padx=10, pady=(2, 8))

            self.output_text = ScrolledText(
                output_card,
                wrap="word",
                bg=self.colors["output_bg"],
                fg=self.colors["output_fg"],
                insertbackground=self.colors["output_fg"],
                relief="flat",
                borderwidth=1,
                font=("Consolas", 10),
            )
            self.output_text.grid(row=2, column=0, sticky="nsew")
            self.output_text.tag_configure("meta", foreground="#90cdf4")
            self.output_text.tag_configure("ok", foreground="#8ce99a")
            self.output_text.tag_configure("warn", foreground="#ffd166")
            self.output_text.tag_configure("error", foreground="#ff8787")
            self.output_text.tag_configure("title", foreground="#74c0fc")
            self.output_text.configure(state="disabled")

            status_frame = ttk.Frame(self.root, style="Main.TFrame", padding=(14, 0, 14, 10))
            status_frame.grid(row=2, column=0, sticky="ew")
            status_frame.columnconfigure(0, weight=1)
            ttk.Label(
                status_frame, textvariable=self.status_var, style="Status.TLabel"
            ).grid(row=0, column=0, sticky="w")

        def _sync_hist_all_state(self) -> None:
            if self.show_hist_var.get():
                self.hist_all_check.state(["!disabled"])
            else:
                self.hist_all_check.state(["disabled"])

        def _load_initial_inputs(self, inputs: Sequence[str], recursive_scan: bool) -> None:
            if not inputs:
                return
            try:
                images = collect_images(inputs, recursive_scan)
            except Exception as exc:
                self.status_var.set(f"输入错误: {exc}")
                messagebox.showerror("输入错误", str(exc))
                return
            self._add_paths([str(path) for path in images])

        def _append_output(self, text: str, tag: str = "") -> None:
            self.output_text.configure(state="normal")
            if tag:
                self.output_text.insert("end", text, tag)
            else:
                self.output_text.insert("end", text)
            self.output_text.see("end")
            self.output_text.configure(state="disabled")

        def _set_result_status(self, result: str, detail: str = "") -> None:
            if result == "PASS":
                display_text = "pass"
            elif result == "FAIL":
                display_text = "fail"
            else:
                display_text = result
            self.result_var.set(display_text)
            self.result_detail_var.set(detail)

            if result == "PASS":
                bg = "#e6fcf5"
                fg = self.colors["success"]
                border = "#b2f2bb"
            elif result == "FAIL":
                bg = "#fff4e6"
                fg = self.colors["danger"]
                border = "#ffd8a8"
            elif "分析中" in result:
                bg = "#e7f5ff"
                fg = self.colors["accent"]
                border = "#a5d8ff"
            else:
                bg = self.colors["result_bg"]
                fg = self.colors["text"]
                border = self.colors["result_border"]

            self.result_box.configure(bg=bg, highlightbackground=border)
            self.result_title.configure(bg=bg, fg=self.colors["muted"])
            self.result_value.configure(bg=bg, fg=fg)
            self.result_detail.configure(bg=bg, fg=self.colors["muted"])

        def _append_stats_block_cn(
            self, stats: Dict[str, object], show_hist: bool, hist_all: bool
        ) -> None:
            block = format_text_cn(stats, show_hist=show_hist, hist_all=hist_all)
            result_tag = "ok" if str(stats.get("test_result")) == "PASS" else "error"
            for line in block.splitlines():
                if line.startswith("测试结果:"):
                    self._append_output(line + "\n", result_tag)
                else:
                    self._append_output(line + "\n")
            self._append_output("\n")

        def _clear_preview(self, message: str = "暂无预览") -> None:
            self.preview_photo = None
            self.preview_label.configure(image="", text=message, fg=self.colors["muted"])
            self.preview_meta_var.set("请选择左侧图片查看缩略图。")

        def _show_preview(self, image_path: str) -> None:
            path = Path(image_path)
            if self.image_tk_module is None:
                self._clear_preview("当前环境不支持缩略图预览")
                self.preview_meta_var.set(path.name)
                return

            try:
                with Image.open(path) as img:
                    width, height = img.size
                    preview = img.convert("RGBA")
                    if hasattr(Image, "Resampling"):
                        preview.thumbnail((340, 220), Image.Resampling.LANCZOS)
                    else:
                        preview.thumbnail((340, 220), Image.LANCZOS)
                    self.preview_photo = self.image_tk_module.PhotoImage(preview)
                    self.preview_label.configure(image=self.preview_photo, text="")
                    self.preview_meta_var.set(f"{path.name}   原图尺寸: {width}x{height}")
            except Exception as exc:
                self._clear_preview(f"预览失败: {exc}")
                self.preview_meta_var.set(path.name)

        def _on_file_selection(self, _event=None) -> None:
            indexes = self.file_list.curselection()
            if not indexes:
                self._clear_preview("暂无预览")
                return
            selected_path = self.selected_paths[indexes[0]]
            self._show_preview(selected_path)
            cached = self.analysis_cache.get(selected_path)
            if cached is None:
                self._set_result_status("未分析", f"当前: {Path(selected_path).name}")
            else:
                self._set_result_status(
                    str(cached.get("test_result", "FAIL")),
                    f"当前: {Path(selected_path).name}",
                )

        def _add_paths(self, raw_paths: Sequence[str]) -> None:
            added = 0
            invalid = 0
            duplicate = 0
            last_added_index = None
            for raw in raw_paths:
                path = Path(raw)
                if not path.is_file() or path.suffix.lower() not in DEFAULT_EXTENSIONS:
                    invalid += 1
                    continue
                resolved = str(path.resolve())
                if resolved in self.selected_set:
                    duplicate += 1
                    continue
                self.selected_set.add(resolved)
                self.selected_paths.append(resolved)
                self.file_list.insert("end", resolved)
                last_added_index = self.file_list.size() - 1
                added += 1

            if added:
                msg = f"新增 {added} 张图片，当前共 {len(self.selected_paths)} 张。"
                if duplicate:
                    msg += f" 跳过重复文件: {duplicate}。"
                if invalid:
                    msg += f" 跳过无效文件: {invalid}。"
                self.status_var.set(msg)
            elif duplicate and not invalid:
                self.status_var.set("所选文件都已在列表中。")
            elif invalid and not duplicate:
                self.status_var.set("未选择到有效图片文件。")
            elif invalid and duplicate:
                self.status_var.set("没有新增文件（无效文件和重复文件已跳过）。")

            if last_added_index is not None:
                self.file_list.selection_clear(0, "end")
                self.file_list.selection_set(last_added_index)
                self.file_list.see(last_added_index)
                self._on_file_selection()

        def select_images(self) -> None:
            patterns = " ".join(f"*{ext}" for ext in sorted(DEFAULT_EXTENSIONS))
            file_paths = filedialog.askopenfilenames(
                title="选择图片文件",
                filetypes=[
                    ("图片文件", patterns),
                    ("所有文件", "*.*"),
                ],
            )
            if not file_paths:
                self.status_var.set("已取消选择。")
                return
            self._add_paths(file_paths)

        def remove_selected_items(self) -> None:
            indexes = list(self.file_list.curselection())
            if not indexes:
                self.status_var.set("未选择要移除的文件。")
                return

            removed = 0
            for index in reversed(indexes):
                path = self.selected_paths.pop(index)
                self.selected_set.discard(path)
                self.analysis_cache.pop(path, None)
                self.file_list.delete(index)
                removed += 1

            self.status_var.set(
                f"已移除 {removed} 个文件，剩余 {len(self.selected_paths)} 个。"
            )
            self._on_file_selection()

        def clear_files(self) -> None:
            self.selected_paths.clear()
            self.selected_set.clear()
            self.analysis_cache.clear()
            self.file_list.delete(0, "end")
            self.status_var.set("已清空文件列表。")
            self._clear_preview()
            self._set_result_status("未分析", "等待分析")

        def clear_output(self) -> None:
            self.output_text.configure(state="normal")
            self.output_text.delete("1.0", "end")
            self.output_text.configure(state="disabled")
            self.status_var.set("已清空结果区。")
            self._set_result_status("未分析", "等待分析")

        def analyze_selected(self) -> None:
            if not self.selected_paths:
                self.status_var.set("请先选择图片。")
                messagebox.showinfo("提示", "请至少选择一张图片文件。")
                return

            selected_indexes = list(self.file_list.curselection())
            if selected_indexes:
                target_paths = [self.selected_paths[i] for i in selected_indexes]
            else:
                target_paths = list(self.selected_paths)

            ignore_transparent_pixels = self.ignore_transparent_var.get()
            include_hist = self.show_hist_var.get()
            include_all_bins = self.hist_all_var.get() if include_hist else False

            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self._append_output("\n" + "=" * 90 + "\n", "meta")
            self._append_output(f"执行时间: {timestamp}\n", "meta")
            self._append_output(
                "参数: "
                f"ignore_transparent={ignore_transparent_pixels}, "
                f"ignore_pure_black=True, "
                "gray_rule=max(R,G,B), "
                f"show_hist={include_hist}, "
                f"hist_all={include_all_bins}\n\n",
                "meta",
            )

            success = 0
            failure = 0
            total = len(target_paths)
            stats_batch: List[Dict[str, object]] = []

            self._set_result_status("分析中", "请稍候")
            self.root.update_idletasks()

            for index, image_path in enumerate(target_paths, start=1):
                self._append_output(f"[{index}/{total}] {image_path}\n", "title")
                try:
                    stats = analyze_image(Path(image_path), ignore_transparent_pixels)
                    self._append_stats_block_cn(
                        stats, show_hist=include_hist, hist_all=include_all_bins
                    )
                    self.analysis_cache[image_path] = stats
                    self._set_result_status(
                        str(stats.get("test_result", "FAIL")),
                        f"当前: {Path(image_path).name}",
                    )
                    stats_batch.append(stats)
                    success += 1
                except Exception as exc:
                    self._append_output(f"失败: {exc}\n\n", "error")
                    self._set_result_status("FAIL", f"异常: {Path(image_path).name}")
                    failure += 1

            if failure:
                self.status_var.set(
                    f"分析完成（有异常）。成功: {success}，失败: {failure}。"
                )
                self._append_output(
                    f"汇总: 成功={success}, 失败={failure}\n", "warn"
                )
            else:
                self.status_var.set(f"分析完成。共处理 {success} 张图片。")
                self._append_output(f"汇总: 成功={success}, 失败=0\n", "ok")

            if not stats_batch and failure:
                self._set_result_status("FAIL", f"异常: {failure}")

            excel_message = _export_results_to_excel(
                stats_batch, _get_excel_output_path(), timestamp
            )
            excel_tag = "error" if "失败" in excel_message else "meta"
            self._append_output(excel_message + "\n", excel_tag)

    try:
        root = tk.Tk()
    except Exception as exc:
        print(f"GUI不可用: {exc}", file=sys.stderr)
        return 2

    GrayCheckApp(root)
    root.mainloop()
    return 0


def main() -> int:
    args = parse_args()

    if args.gui or not args.inputs:
        _hide_console_window()
        return run_gui(
            initial_inputs=args.inputs,
            recursive=args.recursive,
            ignore_transparent=args.ignore_transparent,
            show_hist=args.hist,
            hist_all=args.hist_all,
        )

    try:
        images = collect_images(args.inputs, args.recursive)
    except Exception as exc:
        print(f"Input error: {exc}", file=sys.stderr)
        return 2

    if not images:
        print("No images found.", file=sys.stderr)
        return 2

    results: List[Dict[str, object]] = []
    has_error = False

    for image_path in images:
        try:
            results.append(analyze_image(image_path, args.ignore_transparent))
        except Exception as exc:
            has_error = True
            print(f"Failed: {image_path} ({exc})", file=sys.stderr)

    if not results:
        return 1

    if args.json:
        payload = []
        for item in results:
            if not args.hist:
                copied = dict(item)
                copied.pop("histogram", None)
                payload.append(copied)
            else:
                payload.append(item)
        print(json.dumps(payload, ensure_ascii=False, indent=2))
    else:
        for index, item in enumerate(results):
            if index > 0:
                print()
            print(format_text(item, show_hist=args.hist, hist_all=args.hist_all))

    return 1 if has_error else 0


if __name__ == "__main__":
    sys.exit(main())
